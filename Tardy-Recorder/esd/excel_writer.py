import logging
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook, load_workbook

from .models import Student, TardyEvent
from .utils import week_key, ensure_dirs

logger = logging.getLogger(__name__)


class ExcelWriter:
    def __init__(self, cfg: dict):
        self.cfg = cfg
        self.output_dir = Path(cfg["app"].get("output_dir", "data/output"))
        self.sheet_name = cfg["app"].get("excel_sheet_name", "Tardies")
        self.filename_pattern = cfg["excel"]["filename_pattern"]
        ensure_dirs(self.output_dir)

    def _file_for_event(self, dt: datetime) -> Path:
        wk = week_key(dt)
        name = self.filename_pattern.replace("{YYYY}", str(dt.year)).replace("{WW}", wk)
        return self.output_dir / name

    def _open_or_create(self, path: Path):
        if path.exists():
            wb = load_workbook(path)
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.create_sheet(self.sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            headers = self.cfg["excel"]["headers"]
            ws.append(headers)
        return wb, ws

    def append(self, student: Student, grade: str, event: TardyEvent):
        path = self._file_for_event(event.occurred_at)
        wb, ws = self._open_or_create(path)
        timestamp_str = event.occurred_at.astimezone().strftime("%H:%M:%S")
        date_str = event.occurred_at.astimezone().strftime("%Y-%m-%d")
        row = [student.name, student.id, grade or "", timestamp_str, date_str]
        ws.append(row)
        wb.save(path)
        logger.info("Wrote row to %s: %s", path.name, row)
